#!/usr/bin/env python3
"""
Excel structure analyzer for NanoClaw medical PPT generation.
Summarizes Excel file structure and key data.
Usage: python3 excel_reader.py <path_to_excel>
"""

import sys
import openpyxl


def summarize_excel(path: str) -> str:
    """Summarize Excel file structure and content."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        summary = []

        summary.append(f"File: {path}")
        summary.append(f"Sheet count: {len(wb.sheetnames)}")
        summary.append(f"Sheets: {', '.join(wb.sheetnames)}")
        summary.append("")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            summary.append(f"--- Sheet: {sheet_name} ---")

            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            summary.append(f"Dimensions: {max_row} rows x {max_col} columns")

            # Extract headers (first row with data)
            headers = []
            for col in range(1, min(max_col + 1, 20)):  # Limit to 20 columns
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    break

            if headers:
                summary.append(f"Headers: {', '.join(headers)}")

            # Sample first few data rows
            data_rows = min(max_row - 1, 5)  # Up to 5 data rows
            if data_rows > 0:
                summary.append(f"Data preview (first {data_rows} rows):")
                for row in range(2, min(7, max_row + 1)):
                    row_data = []
                    for col in range(1, len(headers) + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value else "")
                    if any(row_data):  # Only show non-empty rows
                        summary.append(f"  Row {row}: {', '.join(row_data)}")

            summary.append("")

        return "\n".join(summary)

    except Exception as e:
        return f"ERROR: Failed to read Excel file: {e}"


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 excel_reader.py <path_to_excel>")
        sys.exit(1)

    result = summarize_excel(sys.argv[1])
    print(result)
